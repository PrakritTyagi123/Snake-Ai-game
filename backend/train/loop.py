# backend/train/loop.py
from __future__ import annotations

import asyncio
import queue
import threading
import time
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import torch
from openpyxl import Workbook, load_workbook  # for episode logging

# ---------------- Project imports (robust) ----------------
# Prefer package-relative imports (backend is a package). Fall back if needed.
try:
    from .. import config as cfg  # type: ignore
except Exception:  # pragma: no cover
    import config as cfg  # type: ignore

try:
    from ..env.snake_env import SnakeEnv  # type: ignore
    from ..agent.ddqn import DDQNAgent  # type: ignore
    from ..agent.replay_buffer import ReplayBuffer  # type: ignore
except Exception:  # pragma: no cover
    from env.snake_env import SnakeEnv  # type: ignore
    from agent.ddqn import DDQNAgent  # type: ignore
    from agent.replay_buffer import ReplayBuffer  # type: ignore

try:
    from ..llm import LLMReporter  # type: ignore
except Exception:  # pragma: no cover
    try:
        from llm import LLMReporter  # type: ignore
    except Exception:
        LLMReporter = None  # type: ignore


# ---------------- Tunables / constants ----------------
VISUAL_STRIDE = 3       # send a state_frame every N steps (reduce WS/DOM load)
NN_STRIDE = 20          # send nn_activations every N steps
LLM_QUEUE_MAX = 8       # max pending commentary snapshots

EPISODES_HEADERS = [
    "episode", "outcome", "reason", "score", "snake_length", "epsilon",
    "avg_loss", "steps", "duration_ms", "timestamp", "algorithm", "grid_w", "grid_h"
]


def _device() -> str:
    # Only require CUDA if explicitly requested via config
    if getattr(cfg, "REQUIRE_CUDA", 0) and not torch.cuda.is_available():
        raise RuntimeError(
            "CUDA GPU required but not available. "
            "Set REQUIRE_CUDA=0 in config.py (or env) to bypass."
        )
    return "cuda" if torch.cuda.is_available() else "cpu"


def _algo_filename(s: str) -> str:
    return f"{(s or 'ddqn').strip().lower()}.ckpt"


@dataclass
class EpsilonCfg:
    start: float = getattr(cfg, "DEFAULT_EPS_START", 1.0)
    min: float = getattr(cfg, "DEFAULT_EPS_MIN", 0.05)
    decay: float = getattr(cfg, "DEFAULT_EPS_DECAY", 0.995)


@dataclass
class LLMConfig:
    enabled: bool = bool(getattr(cfg, "LLM_ENABLED", 1))
    freq: int = int(getattr(cfg, "LLM_COMMENT_FREQ", 1000))
    backend: str = str(getattr(cfg, "LLM_BACKEND", "hf")).lower()
    model_id: str = str(getattr(cfg, "LLM_MODEL_HF_ID", "Qwen/Qwen2-0.5B-Instruct"))
    local_dir: str = str(getattr(cfg, "LLM_LOCAL_DIR", "ai_chatmodel/Qwen2-0.5B-Instruct"))
    max_tokens: int = int(getattr(cfg, "LLM_MAX_TOKENS", 250))


@dataclass
class TrainerParams:
    # Algo / training
    algorithm: str = "DDQN"
    gamma: float = getattr(cfg, "DEFAULT_GAMMA", 0.95)
    lr: float = getattr(cfg, "DEFAULT_LR", 1e-3)
    batch_size: int = getattr(cfg, "DEFAULT_BATCH", 64)
    max_episodes: int = getattr(cfg, "DEFAULT_MAX_EPISODES", 0)  # 0 => endless
    max_steps: int = getattr(cfg, "DEFAULT_MAX_STEPS", 1000)

    # Board
    board: Dict[str, int] = field(
        default_factory=lambda: {
            "w": getattr(cfg, "BOARD_WIDTH", 20),
            "h": getattr(cfg, "BOARD_HEIGHT", 20),
        }
    )
    win_length: int = getattr(
        cfg,
        "WIN_LENGTH",
        (getattr(cfg, "BOARD_WIDTH", 20) * getattr(cfg, "BOARD_HEIGHT", 20)) - 1,
    )

    # Live controls
    step_delay_ms: int = getattr(cfg, "STEP_DELAY_MS", 0)
    epsilon: EpsilonCfg = field(default_factory=EpsilonCfg)

    # Misc
    target_update_freq: int = getattr(cfg, "TARGET_UPDATE_FREQ", 250)
    llm: LLMConfig = field(default_factory=LLMConfig)


class Trainer:
    """
    Full trainer:
      - streams frames/metrics over a broadcaster
      - runs DDQN training
      - NO autosave; saves only on pause/close or explicit /api/save
      - single checkpoint file per algo: storage/checkpoints/<algo>.ckpt
      - episode summaries are appended to storage/episodes/<algo>.xlsx
    """
    def __init__(self, async_broadcast: Callable[[Dict[str, Any]], Any]):
        self.params = TrainerParams()
        self._broadcast_async = async_broadcast

        # Async loop to drive the (possibly) async broadcaster from a background thread
        self._loop = asyncio.new_event_loop()
        self._loop_th = threading.Thread(target=self._run_loop, daemon=True)
        self._loop_th.start()

        # Runtime flags
        self._th: Optional[threading.Thread] = None
        self._stop_evt = threading.Event()
        self._pause_evt = threading.Event()
        self._lock = threading.Lock()
        self.closing = False  # when True, stop emitting late WS messages

        # Stats
        self.total_steps: int = 0
        self.episode: int = 0
        self._eps_value: float = self.params.epsilon.start

        # RL components
        self.device = _device()
        self.env = SnakeEnv(self.params.board["w"], self.params.board["h"], self.params.win_length)
        state = self.env.reset()
        self.state_dim = len(state)
        self.agent = DDQNAgent(self.state_dim, out_dim=4, lr=self.params.lr, gamma=self.params.gamma, device=self.device)
        self.buffer = ReplayBuffer(capacity=100_000, seed=0)

        # Episode logging buffer
        self._episodes_buf: List[Dict[str, Any]] = []
        self._episodes_lock = threading.Lock()

        # --- LLM (non-blocking worker), hard-disable aware ---
        hard_off = bool(getattr(cfg, "LLM_HARD_DISABLE", 0))
        if hard_off:
            try:
                self.params.llm.enabled = False
            except Exception:
                pass

        self._llm: Optional[Any] = None  # keep typing simple even if llm import fails
        self._llm_last_step = 0
        self._llm_q: "queue.Queue[Optional[Dict[str, Any]]]" = queue.Queue(maxsize=LLM_QUEUE_MAX)
        self._llm_stop_evt = threading.Event()
        self._llm_th: Optional[threading.Thread] = None

        # start worker only if enabled
        if bool(self.params.llm.enabled):
            self._llm_th = threading.Thread(target=self._llm_worker, daemon=True)
            self._llm_th.start()

        self._maybe_init_llm()

    # ---------- asyncio bridge ----------
    def _run_loop(self):
        asyncio.set_event_loop(self._loop)
        self._loop.run_forever()

    def _emit(self, payload: Dict[str, Any]) -> None:
        if self.closing:
            return
        try:
            fut = self._broadcast_async(payload)
            if asyncio.iscoroutine(fut):
                asyncio.run_coroutine_threadsafe(fut, self._loop)
        except Exception:
            pass

    # ---------- config ----------
    def get_config(self) -> Dict[str, Any]:
        p = asdict(self.params)
        p["epsilon"] = asdict(self.params.epsilon)
        p["llm"] = asdict(self.params.llm)
        return p

    def update_config(self, patch: Dict[str, Any]) -> Dict[str, Any]:
        if not patch:
            return {"ok": True, "params": self.get_config()}
        with self._lock:
            # Track changes that need live-application to the agent
            lr_changed = ("lr" in patch and patch["lr"] is not None)
            gamma_changed = ("gamma" in patch and patch["gamma"] is not None)
            eps_start_changed = False
            eps_min_changed = False

            # Simple scalar fields copied into params
            for k in ("algorithm", "lr", "gamma", "batch_size", "max_episodes",
                      "max_steps", "step_delay_ms", "target_update_freq"):
                if k in patch and patch[k] is not None:
                    setattr(self.params, k, patch[k])

            # Epsilon nested
            if "epsilon" in patch and isinstance(patch["epsilon"], dict):
                ep = patch["epsilon"]
                if "start" in ep and ep["start"] is not None:
                    setattr(self.params.epsilon, "start", float(ep["start"]))
                    eps_start_changed = True
                if "min" in ep and ep["min"] is not None:
                    setattr(self.params.epsilon, "min", float(ep["min"]))
                    eps_min_changed = True
                if "decay" in ep and ep["decay"] is not None:
                    setattr(self.params.epsilon, "decay", float(ep["decay"]))

                # REAL-TIME epsilon behavior:
                if eps_start_changed:
                    self._eps_value = float(self.params.epsilon.start)
                if eps_min_changed:
                    self._eps_value = max(float(self.params.epsilon.min), float(self._eps_value))

            # Board change → rebuild env (next frame)
            if "board" in patch and isinstance(patch["board"], dict):
                w = int(patch["board"].get("w", self.params.board["w"]))
                h = int(patch["board"].get("h", self.params.board["h"]))
                if w != self.params.board["w"] or h != self.params.board["h"]:
                    self.params.board["w"] = w
                    self.params.board["h"] = h
                    self.params.win_length = w * h - 1
                    self.env = SnakeEnv(w, h, self.params.win_length)
                    self._emit({"type": "board_changed", "board": {"w": w, "h": h}})

            # ---- LIVE-APPLY LR and GAMMA to the agent immediately ----
            if lr_changed and hasattr(self, "agent") and hasattr(self.agent, "optimizer"):
                try:
                    new_lr = float(self.params.lr)
                    for g in self.agent.optimizer.param_groups:
                        g["lr"] = new_lr
                except Exception:
                    pass

            if gamma_changed and hasattr(self, "agent"):
                try:
                    setattr(self.agent, "gamma", float(self.params.gamma))
                except Exception:
                    pass

            # LLM changes (respect hard-disable, start/stop worker on demand)
            if "llm" in patch and isinstance(patch["llm"], dict):
                if bool(getattr(cfg, "LLM_HARD_DISABLE", 0)):
                    try:
                        setattr(self.params.llm, "enabled", False)
                    except Exception:
                        pass
                else:
                    for lk in ("enabled", "freq", "backend", "model_id", "local_dir", "max_tokens"):
                        if lk in patch["llm"]:
                            setattr(self.params.llm, lk, patch["llm"][lk])
                    # start/stop worker if toggled
                    try:
                        if self.params.llm.enabled and self._llm_th is None:
                            self._llm_th = threading.Thread(target=self._llm_worker, daemon=True)
                            self._llm_th.start()
                        elif (not self.params.llm.enabled) and self._llm_th:
                            self._llm_stop_evt.set()
                            try:
                                self._llm_q.put_nowait(None)
                            except Exception:
                                pass
                            if self._llm_th.is_alive():
                                self._llm_th.join(timeout=1.0)
                            self._llm_th = None
                    except Exception:
                        pass
                self._maybe_init_llm()

        # Let UI know the currently applied params (and live epsilon value)
        self._emit({"type": "config_applied", "params": self.get_config(), "epsilon_current": self._eps_value})
        return {"ok": True, "params": self.get_config()}

    def _maybe_init_llm(self):
        # Respect hard disable and availability
        if bool(getattr(cfg, "LLM_HARD_DISABLE", 0)):
            self._llm = None
            return
        if not LLMReporter or not bool(self.params.llm.enabled):
            self._llm = None
            return
        try:
            self._llm = LLMReporter(
                backend=self.params.llm.backend,
                model_id=self.params.llm.model_id,
                local_dir=self.params.llm.local_dir,
                max_tokens=self.params.llm.max_tokens
            )
        except Exception:
            self._llm = None

    # ---------- LLM worker ----------
    def _llm_worker(self):
        """Background thread: turns snapshots into commentary without blocking training."""
        while not self._llm_stop_evt.is_set():
            try:
                item = self._llm_q.get(timeout=0.25)
            except queue.Empty:
                continue
            if item is None:
                break  # sentinel
            if self.closing or not self._llm:
                continue
            try:
                text = self._llm.generate_comment(item)
                if text:
                    self._emit({
                        "type": "llm_comment",
                        "episode": item.get("episode", 0),
                        "step": item.get("step", 0),
                        "text": text
                    })
            except Exception:
                # commentary must never break training
                pass

    # ---------- lifecycle ----------
    def start(self) -> None:
        if self._th and self._th.is_alive():
            return
        self._stop_evt.clear()
        self._pause_evt.clear()
        self.closing = False
        self._th = threading.Thread(target=self._train_loop, daemon=True)
        self._th.start()

    def resume(self) -> None:
        self._pause_evt.clear()
        self._emit({"type": "metrics_tick", "episode": self.episode, "step": self.total_steps})

    def stop(self) -> None:
        self._stop_evt.set()

    def reset(self) -> None:
        self.total_steps = 0
        self.episode = 0
        self._eps_value = self.params.epsilon.start
        self.env = SnakeEnv(self.params.board["w"], self.params.board["h"], self.params.win_length)
        self.env.reset()
        self.buffer = ReplayBuffer(capacity=100_000, seed=0)
        self._emit({"type": "config_applied", "params": self.get_config()})

    async def pause_and_save(self) -> None:
        self._pause_evt.set()
        await asyncio.sleep(0.01)  # let loop yield
        await self.save_checkpoint(reason="pause")

    def shutdown(self) -> None:
        """Stop training thread, LLM worker, flush logs, and internal asyncio loop cleanly."""
        # stop training
        try:
            self._stop_evt.set()
            self._pause_evt.set()
            if self._th and self._th.is_alive():
                self._th.join(timeout=1.0)
        except Exception:
            pass
        # flush episode log
        try:
            self.flush_episode_log()
        except Exception:
            pass
        # stop LLM worker
        try:
            self._llm_stop_evt.set()
            try:
                self._llm_q.put_nowait(None)  # wake the worker
            except Exception:
                pass
            if self._llm_th and self._llm_th.is_alive():
                self._llm_th.join(timeout=1.0)
        except Exception:
            pass
        # stop broadcaster loop
        try:
            if self._loop.is_running():
                self._loop.call_soon_threadsafe(self._loop.stop)
            if self._loop_th and self._loop_th.is_alive():
                self._loop_th.join(timeout=1.0)
        except Exception:
            pass

    # ---------- checkpointing ----------
    def _ckpt_dir(self) -> Path:
        d = Path(getattr(cfg, "CHECKPOINT_DIR", "storage/checkpoints"))
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _ckpt_path(self) -> Path:
        return self._ckpt_dir() / _algo_filename(self.params.algorithm)

    async def save_checkpoint(self, reason: str = "manual") -> None:
        self._emit({"type": "saving_started", "reason": reason})
        path = self._ckpt_path()
        tmp = path.with_suffix(".tmp")
        try:
            ckpt: Dict[str, Any] = {
                "version": 1,
                "params": self.get_config(),
                "episode": self.episode,
                "total_steps": self.total_steps,
                "eps_value": self._eps_value,
                "agent": {
                    "policy": self.agent.policy.state_dict(),
                    "target": self.agent.target.state_dict(),
                    "opt": self.agent.optimizer.state_dict(),
                },
                "replay": self.buffer.state_dict(),
            }
            torch.save(ckpt, tmp.as_posix())
            tmp.replace(path)  # atomic-ish on Windows/Posix
            self._emit({"type": "saving_finished", "path": path.as_posix()})
        except Exception as e:
            try:
                if tmp.exists():
                    tmp.unlink(missing_ok=True)
            except Exception:
                pass
            self._emit({"type": "saving_failed", "error": str(e)})

    def load_checkpoint(self) -> bool:
        path = self._ckpt_path()
        if not path.exists():
            return False
        try:
            ckpt = torch.load(path.as_posix(), map_location=self.device)
            p = ckpt.get("params") or {}
            self.update_config(p)
            self.episode = int(ckpt.get("episode") or 0)
            self.total_steps = int(ckpt.get("total_steps") or 0)
            self._eps_value = float(ckpt.get("eps_value") or self.params.epsilon.start)

            agent_sd = ckpt.get("agent") or {}
            if agent_sd:
                self.agent.policy.load_state_dict(agent_sd["policy"])
                self.agent.target.load_state_dict(agent_sd["target"])
                self.agent.optimizer.load_state_dict(agent_sd["opt"])

            rb_sd = ckpt.get("replay")
            if rb_sd:
                self.buffer.load_state_dict(rb_sd)

            # notify UI that a checkpoint was restored
            self._emit({
                "type": "checkpoint_loaded",
                "path": path.as_posix(),
                "episode": self.episode,
                "total_steps": self.total_steps
            })
            self._emit({"type": "config_applied", "params": self.get_config()})
            return True
        except Exception:
            return False

    # ---------- episode logging (Excel) ----------
    def _episodes_dir(self) -> Path:
        d = Path(getattr(cfg, "EPISODES_DIR", "storage/episodes"))
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _episodes_path(self) -> Path:
        algo = (self.params.algorithm or "ddqn").strip().lower()
        return self._episodes_dir() / f"{algo}.xlsx"

    def _ensure_wb_and_sheet(self, path: Path):
        if path.exists():
            wb = load_workbook(path.as_posix())
            ws = wb.active
            if ws.max_row < 1:
                ws.append(EPISODES_HEADERS)
            return wb, ws
        wb = Workbook()
        ws = wb.active
        ws.title = "episodes"
        ws.append(EPISODES_HEADERS)
        return wb, ws

    def _log_episode(self, row: Dict[str, Any]) -> None:
        rec = dict(row)
        rec["algorithm"] = self.params.algorithm
        rec["grid_w"] = self.params.board["w"]
        rec["grid_h"] = self.params.board["h"]
        with self._episodes_lock:
            self._episodes_buf.append(rec)
            batch = int(getattr(cfg, "EPISODES_WRITE_BATCH", 100))
            if len(self._episodes_buf) >= max(1, batch):
                try:
                    self._flush_episode_log_locked()
                except Exception:
                    pass  # never interrupt training

    def flush_episode_log(self) -> None:
        """Public flush (call on close/shutdown)."""
        with self._episodes_lock:
            self._flush_episode_log_locked()

    def _flush_episode_log_locked(self) -> None:
        if not self._episodes_buf:
            return
        path = self._episodes_path()
        wb, ws = self._ensure_wb_and_sheet(path)
        buf = self._episodes_buf
        self._episodes_buf = []
        for r in buf:
            ws.append([r.get(k, "") for k in EPISODES_HEADERS])
        wb.save(path.as_posix())

    # ---------- training loop ----------
    def _train_loop(self):
        state = self.env.reset()
        ep_start_ms = int(time.time() * 1000)
        last_q: Optional[List[float]] = None
        loss_accum: float = 0.0
        loss_count: int = 0
        steps_in_ep: int = 0

        while not self._stop_evt.is_set():
            if self._pause_evt.is_set():
                time.sleep(0.02)
                continue

            # ε-greedy
            action = self.agent.act(state, self._eps_value)
            next_state, reward, done, info = self.env.step(action)
            self.buffer.push(state, action, reward, next_state, float(done))

            # learn
            if len(self.buffer) >= self.params.batch_size:
                batch = self.buffer.sample(self.params.batch_size)
                update_target = (self.total_steps % max(1, int(self.params.target_update_freq))) == 0
                loss = self.agent.learn(batch, target_net_update=update_target)
                try:
                    loss_val = float(loss.detach().item()) if hasattr(loss, "detach") else float(loss)
                except Exception:
                    loss_val = float(loss) if isinstance(loss, (int, float)) else 0.0
                loss_accum += loss_val
                loss_count += 1

            # q-values for viz
            try:
                with torch.no_grad():
                    s = torch.tensor(state, dtype=torch.float32, device=self.agent.device).unsqueeze(0)
                    q = self.agent.policy(s)[0].detach().float().cpu().tolist()
                    last_q = q
            except Exception:
                last_q = None

            # counters
            self.total_steps += 1
            steps_in_ep += 1
            self._eps_value = max(self.params.epsilon.min, self._eps_value * float(self.params.epsilon.decay))

            # stream frame (decimated)
            if (self.total_steps % VISUAL_STRIDE) == 0:
                self._emit({"type": "state_frame", "grid": self.env.render_spec()})

            # stream activations occasionally (decimated, include sizes)
            if (self.total_steps % NN_STRIDE) == 0:
                layers_payload: List[Dict[str, Any]] = []
                sizes: List[int] = []
                try:
                    act = self.agent.activation_summary()  # dict[str, list[float]] or similar
                except Exception:
                    act = None
                if act:
                    for _, values in act.items():
                        v = list(values) if values is not None else []
                        layers_payload.append({"act": v})
                        sizes.append(len(v))
                if last_q is not None:
                    layers_payload.append({"q": last_q})
                    out_size = len(last_q)
                else:
                    out_size = 4
                if layers_payload:
                    self._emit({
                        "type": "nn_activations",
                        "layers": layers_payload,
                        "sizes": {"hidden": sizes, "output": out_size}
                    })

            # metrics tick
            if (self.total_steps % 5) == 0:
                self._emit({"type": "metrics_tick", "episode": self.episode, "step": self.total_steps})

            # non-blocking LLM commentary
            if self.params.llm.freq > 0 and (self.total_steps - self._llm_last_step) >= self.params.llm.freq:
                self._llm_last_step = self.total_steps
                snap = {
                    "step": self.total_steps,
                    "episode": self.episode,
                    "epsilon": self._eps_value,
                    "avg_loss": (loss_accum / max(1, loss_count)),
                    "score": self.env.score,
                }
                try:
                    self._llm_q.put_nowait(snap)
                except queue.Full:
                    pass

            # optional delay for visuals
            d = max(0, int(self.params.step_delay_ms))
            if d:
                time.sleep(d / 1000.0)

            # episode end
            if done or (self.params.max_steps > 0 and steps_in_ep >= self.params.max_steps):
                outcome = (info.get("outcome") if isinstance(info, dict) else None) or (
                    "WIN" if len(self.env.snake) >= self.params.win_length else "LOSS"
                )
                reason = (info.get("reason") if isinstance(info, dict) else outcome)
                avg_loss = (loss_accum / max(1, loss_count))
                row = {
                    "episode": self.episode,
                    "outcome": outcome,
                    "reason": reason,
                    "score": self.env.score,
                    "snake_length": len(self.env.snake),
                    "epsilon": self._eps_value,
                    "avg_loss": avg_loss,
                    "steps": steps_in_ep,
                    "duration_ms": int(time.time() * 1000) - ep_start_ms,
                    "timestamp": int(time.time() * 1000),
                }
                self._emit({"type": "episode_summary", **row})
                self._log_episode(row)

                if outcome == "WIN":
                    self._emit({"type": "game_over", "episode": self.episode, "outcome": "WIN",
                                "win_length": len(self.env.snake), "message": "YOU WIN!"})

                # reset episode
                self.episode += 1
                self.env = SnakeEnv(self.params.board["w"], self.params.board["h"], self.params.win_length)
                state = self.env.reset()
                ep_start_ms = int(time.time() * 1000)  # <-- start timer for the new episode
                steps_in_ep = 0
                loss_accum = 0.0
                loss_count = 0
            else:
                state = next_state

        # loop exited
        self._emit({"type": "run_stopped", "episode": self.episode, "step": self.total_steps})
